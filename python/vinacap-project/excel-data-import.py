import openpyxl as px
from openpyxl import *

path = "C:/Users/caoda/Desktop/code/python-stuff/vinacap-project/vinacapital.xlsx"
path2 = "C:/Users/caoda/Desktop/code/python-stuff/vinacap-project/answers.xlsx"
path3 = "C:/Users/caoda/Desktop/code/python-stuff/vinacap-project/green.xlsx"

wb = px.load_workbook(path, data_only=True)
wb2 = px.load_workbook(path2, data_only=True)
wb3 = px.load_workbook(path3, data_only=True)
wb.active = 1
sheet = wb.active
sheet2 = wb2.active
sheet3 = wb3.active

for i in range(451, 600): # name
    s2_name = str('C' + str(i+1))
    rannum = str('B' + str(i+1))
    s1_name = str('D' + str(int(sheet2[rannum].value + 1)))
    index = str('A' + str(int(sheet2[rannum].value + 1)))
    if sheet[index].value == sheet2[rannum].value:
        sheet2[s2_name].value = sheet[s1_name].value

for i in range(451, 600): # hospital
    s2_hos = str('D' + str(i+1))
    rannum = str('B' + str(i+1))
    s1_hos = str('J' + str(int(sheet2[rannum].value + 1)))
    index = str("A" + str(int(sheet2[rannum].value + 1)))
    if sheet[index].value == sheet2[rannum].value:
        sheet2[s2_hos].value = sheet[s1_hos].value

for i in range(451, 600): # year
    s2_year = str('E' + str(i+1))
    rannum = str('B' + str(i+1))
    s1_year = str('L' + str(int(sheet2[rannum].value + 1)))
    index = str("A" + str(int(sheet2[rannum].value + 1)))
    if sheet[index].value == sheet2[rannum].value:
        sheet2[s2_year].value = sheet[s1_year].value

for i in range(451, 600): # tel
    s2_tel = str('F' + str(i+1))
    rannum = str('B' + str(i+1))
    s1_tel = str('H' + str(int(sheet2[rannum].value + 1)))
    index = str("A" + str(int(sheet2[rannum].value + 1)))
    if sheet[index].value == sheet2[rannum].value:
        sheet2[s2_tel].value = sheet[s1_tel].value

# count greens (answered)
s3_row, count = 0, 0
for i in range(1, 600):
    if sheet2['C' + str(i+1)].fill.fgColor.rgb == 'FFd9ead3' or sheet2['C' + str(i+1)].fill.fgColor.rgb == 'FFD9EAD3':
        s3_row += 1
        count += 1
        sheet3['A' + str(s3_row)].value = sheet2['A' + str(i+1)].value
        sheet3['B' + str(s3_row)].value = sheet2['B' + str(i+1)].value
        sheet3['C' + str(s3_row)].value = sheet2['C' + str(i+1)].value
        sheet3['D' + str(s3_row)].value = sheet2['D' + str(i+1)].value
        sheet3['E' + str(s3_row)].value = sheet2['E' + str(i+1)].value
        sheet3['F' + str(s3_row)].value = sheet2['F' + str(i+1)].value
        sheet3['G' + str(s3_row)].value = sheet2['G' + str(i+1)].value
        sheet3['H' + str(s3_row)].value = sheet2['H' + str(i+1)].value
        sheet3['I' + str(s3_row)].value = sheet2['I' + str(i+1)].value
        sheet3['J' + str(s3_row)].value = sheet2['J' + str(i+1)].value
print(count)

# count reds (not answered)
s3_row += 1
for i in range(1, 600):
    if sheet2['C' + str(i+1)].fill.fgColor.rgb == 'FFf4cccc' or sheet2['C' + str(i+1)].fill.fgColor.rgb == 'FFF4CCCC':
        s3_row += 1
        count += 1
        sheet3['A' + str(s3_row)].value = sheet2['A' + str(i+1)].value
        sheet3['B' + str(s3_row)].value = sheet2['B' + str(i+1)].value
        sheet3['C' + str(s3_row)].value = sheet2['C' + str(i+1)].value
        sheet3['D' + str(s3_row)].value = sheet2['D' + str(i+1)].value
        sheet3['E' + str(s3_row)].value = sheet2['E' + str(i+1)].value
        sheet3['F' + str(s3_row)].value = sheet2['F' + str(i+1)].value
        sheet3['G' + str(s3_row)].value = sheet2['G' + str(i+1)].value
        sheet3['H' + str(s3_row)].value = sheet2['H' + str(i+1)].value
        sheet3['I' + str(s3_row)].value = sheet2['I' + str(i+1)].value
        sheet3['J' + str(s3_row)].value = sheet2['J' + str(i+1)].value
print(count)

wb2.save("test.xlsx")
wb3.save("green.xlsx")